import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from tabulate import tabulate
import time


#Arquivo excel
arquivo_excel = r"Local da planilha de solicitção de acessos"
df = pd.read_excel(arquivo_excel)
df.columns = df.columns.astype(str).str.strip()

def navegador (pagina):
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    navegador = webdriver.Chrome(options=options)
    navegador.maximize_window()
    navegador.get(pagina)
    time.sleep(4)
    return navegador

chamados_criados = []
def janela_login (navegador, email_usuario, senha_usuario):
 # Preencher e-mail
    email = navegador.find_element(By.XPATH, '//*[@id="userNameInput"]')
    email.clear()
    email.send_keys(email_usuario)
    
    # Preencher senha
    senha = navegador.find_element(By.XPATH, '//*[@id="passwordInput"]')
    senha.clear()
    senha.send_keys(senha_usuario)
    
    # Botão de login
    navegador.find_element(By.XPATH, '//*[@id="submitButton"]').click()
    time.sleep(3)
    print("Acesso concluido")
def janela_principal_acesso (nome_usuario, atendimento_posto, nivel_hierarquico, gerente_colab, sistema_solicitado, perfil_solicitado, descricao_user):
    try:
        #preencher nome
        nome = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-22"]')
        nome.click()
        time.sleep(3)
        nome = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen22_search"]')
        nome.send_keys(nome_usuario)
        time.sleep(2)
        nome.send_keys(Keys.TAB)
        time.sleep(2)

        #preencher atendimento
        atendimento = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_local_atendimento2"]/a')
        atendimento.click()
        time.sleep(2)
        atendimento = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen3_search"]')
        atendimento.send_keys(atendimento_posto)
        time.sleep(2)
        atendimento.send_keys(Keys.TAB)
        time.sleep(2)

        #preencher nivel hierárquico
        nivel = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-1"]')
        nivel.click()
        time.sleep(2)
        nivel = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen1_search"]')
        nivel.send_keys(nivel_hierarquico)
        time.sleep(2)
        nivel.send_keys(Keys.TAB)
        time.sleep(2)

        #preencher gerente
        gerente = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_select_the_manager"]/a')
        gerente.click()
        time.sleep(2)
        gerente = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen6_search"]')
        gerente.send_keys(gerente_colab)
        time.sleep(2)
        gerente.send_keys(Keys.TAB)
        time.sleep(2)

        #informar sistema
        sistema = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_inform_the_system"]/a')
        sistema.click()
        sistema = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen9_search"]')
        sistema.send_keys(sistema_solicitado)
        time.sleep(2)
        sistema.send_keys(Keys.TAB)
        time.sleep(2)

        #inserir perfil
        perfil = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_enter_your_profile"]/a')
        perfil.click()
        perfil = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen11_search"]')
        perfil.send_keys(perfil_solicitado)
        time.sleep(2)
        perfil.send_keys(Keys.TAB)
        time.sleep(2)

        #descrição
        descricao = navegador.find_element(By.XPATH, '//*[@id="sp_formfield_descreva"]')
        descricao.send_keys(descricao_user)
        time.sleep(6)

        #clicar no botão
        navegador.find_element(By.XPATH, '//*[@id="sc_cat_item"]/div[1]/div[2]/div/div[1]/div[3]/button').click()
        time.sleep(5)

        #numero do chamado
        numero_chamado = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="x69373df21bd58850d827535f1b4bcba1"]/div/div/dl[1]/dd[1]'))
        )
        chamado_numero = numero_chamado.text
        print(f"Número do chamado: {chamado_numero}")
        return chamado_numero 

    except Exception as e:
        print(f"Erro na função janela_principal: {e}")
        return None
def janela_principal_vpn(nome_usuario_vpn, atendimento_posto_vpn, localidade_vpn, acesso_vpn, terceiro_vpn, perfil_solicitado_vpn, email_terceiro_vpn, descricao_terceiro_vpn,descricao_user_vpn, nivel_hierarquico_vpn, gerente_colab_vpn):
    try:
        #preencher nome
        nome = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-14"]')
        nome.click()
        time.sleep(3)
        nome = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen14_search"]')
        nome.send_keys(nome_usuario_vpn)
        time.sleep(2)
        nome.send_keys(Keys.TAB)
        time.sleep(2)

        #preencher atendimento
        atendimento = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_local_atendimento2"]/a')
        atendimento.click()
        time.sleep(2)
        atendimento = navegador.find_element(By.XPATH, '//*[@id="s2id_sp_formfield_local_atendimento2"]/a')
        atendimento.send_keys(atendimento_posto_vpn)
        time.sleep(2)
        atendimento.send_keys(Keys.TAB)
        time.sleep(2)

        #Localidade
        localidade = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-6"]')
        localidade.click()
        localidade = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen6_search"]')
        localidade.send_keys(localidade_vpn)
        time.sleep(2)
        localidade.send_keys(Keys.TAB)
        time.sleep(2)
        
        #Acesso
        acesso = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-6"]')
        acesso.click()
        acesso = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen6_search"]')
        acesso.send_keys(acesso_vpn)
        time.sleep(2)
        acesso.send_keys(Keys.TAB)
        time.sleep(2)
        
        #Terceiro
        terceiro = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-3"]')
        terceiro.click()
        terceiro = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen3_search"]')
        terceiro.send_keys(terceiro_vpn)
        time.sleep(2)
        terceiro.send_keys(Keys.TAB)
        time.sleep(2)

        #inserir perfil
        perfil = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-2"]')
        perfil.click()
        perfil = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen2_search"]')
        perfil.send_keys(perfil_solicitado_vpn)
        time.sleep(2)
        perfil.send_keys(Keys.TAB)
        time.sleep(2)
        
        #Email terceiro
        email_terceiro = navegador.find_element(By.XPATH, '//*[@id="sp_formfield_third_party_email"]')
        email_terceiro.send_keys(email_terceiro_vpn)
        time.sleep(6)
        
        #descrição de atividade terceiro
        descricao_atividade = navegador.find_element(By.XPATH, '//*[@id="sp_formfield_descreva"]')
        descricao_atividade.send_keys(descricao_terceiro_vpn)
        time.sleep(6)
        
        #descrição de atividade próprio
        descricao = navegador.find_element(By.XPATH, '//*[@id="sp_formfield_justification"]')
        descricao.send_keys(descricao_user_vpn)
        time.sleep(6)
        
        #preencher nivel hierárquico
        nivel = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-4"]')
        nivel.click()
        time.sleep(2)
        nivel = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen4_search"]')
        nivel.send_keys(nivel_hierarquico_vpn)
        time.sleep(2)
        nivel.send_keys(Keys.TAB)
        time.sleep(2)

        #preencher gerente
        gerente = navegador.find_element(By.XPATH, '//*[@id="select2-chosen-8"]')
        gerente.click()
        time.sleep(2)
        gerente = navegador.find_element(By.XPATH, '//*[@id="s2id_autogen8_search"]')
        gerente.send_keys(gerente_colab_vpn)
        time.sleep(2)
        gerente.send_keys(Keys.TAB)
        time.sleep(2)       


        #clicar no botão
        navegador.find_element(By.XPATH, '//*[@id="sc_cat_item"]/div[1]/div[2]/div/div[1]/div[3]/button').click()
        time.sleep(5)

        #numero do chamado
        numero_chamado = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="x69373df21bd58850d827535f1b4bcba1"]/div/div/dl[1]/dd[1]'))
        )
        chamado_numero = numero_chamado.text
        print(f"Número do chamado: {chamado_numero}")
        return chamado_numero 

    except Exception as e:
        print(f"Erro na função janela_principal: {e}")
        return None
    
def salvar_chamado(index, chamado_numero, nome_usuario, sistema_solicitado):
    wb = load_workbook(arquivo_excel)
    ws = wb['Chamado']

    ws.cell(row = index + 2, column = 14).value = chamado_numero

    wb.save(arquivo_excel)
    print(f"Chamado {chamado_numero} salvo na planilha, na linha {index + 2}")

    chamados_criados.append([nome_usuario, sistema_solicitado, chamado_numero])


def main():
    #Acessos
    email_usuario = 'Email do criador'
    senha_usuario = 'Senha do criador'
    atendimento_posto = "Informação para criar acesso"
    nivel_hierarquico = "Informação para criar acesso"
    gerente_colab = "Informação para criar acesso"
    perfil_solicitado = "Informação para criar acesso"
    descricao_user = "Informação para criar acesso"
    link_acesso = 'Link do site para criar acessos'
    
    #VPN
    atendimento_posto_vpn = "Informação para criar acesso"
    localidade_vpn = "Informação para criar acesso"
    perfil_solicitado_vpn = "Informação para criar acesso"
    nivel_hierarquico_vpn = "Informação para criar acesso"
    gerente_colab_vpn = "Informação para criar acesso"
    link_vpn = 'Link do site para criar acessos'

    navegador_instance = navegador(link_vpn)

    janela_login(navegador_instance,email_usuario, senha_usuario)
            
    #Buscar dados do excel
    for index, row in df.iterrows():
        print(f"Linha: {index + 1}")
        
        if pd.isna(row['Nome completo']):
            print(f"Pular linha {index + 1} - Nome completo está vazio")
            continue

        if not pd.isna(row['Chamado']):
            print(f"Pular linha {index} - Chamado já realizado")
            continue  # Pula para a próxima linha se já há um valor em "Nº Status"

        # Dados para abrir o chamado

        if row['Coluna onde verificar a solicitação'] == 'Acesso':

            pagina = link_acesso
            nome_usuario = row['Nome completo']
            matricula = row['Matrícula']
            empresa = row['Empresa']
            sistema_solicitado = row['Acesso']
            descricao = row['Atividade no acesso']
            cargo = row['Cargo']
            
            if sistema_solicitado == 'Informação para criar acesso':
                sistema_solicitado = 'Informação para criar acesso'
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"""Informação para criar acesso: {matricula} \nInformação para criar acesso:{empresa} \nInformação para criar acesso: {cargo}
                \nInformação para criar acesso \nInformação para criar acesso \n{descricao}"""
            
            elif sistema_solicitado == 'Informação para criar acesso':
                sistema_solicitado = 'Informação para criar acesso'
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                
            elif sistema_solicitado == 'Informação para criar acesso':
                sistema_solicitado = 'Informação para criar acesso'
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
            
            elif sistema_solicitado == 'Informação para criar acesso':
                sistema_solicitado = 'Informação para criar acesso'
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
            
        
            chamado = janela_principal_acesso(nome_usuario, atendimento_posto, nivel_hierarquico, gerente_colab, sistema_solicitado, perfil_solicitado, descricao_user, pagina)
        
        elif row['Coluna onde verificar a solicitação'] == 'Aplicativo':
            
            pagina = link_acesso
            nome_usuario = row['Nome completo']
            matricula = row['Matrícula']
            empresa = row['Empresa']
            sistema_solicitado = 'Informação para criar acesso'
            perfil_solicitado = row['Acesso']
            descricao = row['Atividade no acesso']
            cargo = row['Cargo']

            if perfil_solicitado == 'Informação para criar acesso':
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"""Informação para criar acesso: {matricula} \nInformação para criar acesso:{empresa} \nInformação para criar acesso: {cargo}
                \nInformação para criar acesso \nInformação para criar acesso
                \nInformação para criar acesso \n{descricao}"""
        
            elif perfil_solicitado == 'Informação para criar acesso':
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                
            elif perfil_solicitado == 'Informação para criar acesso':
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
            
            elif perfil_solicitado == 'Informação para criar acesso':
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                
            elif perfil_solicitado == 'Informação para criar acesso':
                perfil_solicitado = 'Informação para criar acesso'
                descricao_user = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                
            
            chamado = janela_principal_acesso(nome_usuario, atendimento_posto, nivel_hierarquico, gerente_colab, sistema_solicitado, perfil_solicitado, descricao_user, pagina)

        elif row['Coluna onde verificar a solicitação'] == 'VPN':

            pagina = link_vpn
            nome_usuario_vpn = row['Nome completo']
            matricula = row['Matrícula']
            empresa = row['Empresa']
            acesso_vpn = row[acesso]
            terceiro_vpn = parceira
            email_terceiro_vpn = row['Email']
            descricao_terceiro_vpn = row['Atividade no acesso']
            descricao_user_vpn = row['Atividade no acesso']
            
            if pd.isna(row['Tipo de solicitação do VPN']) == 'Solicitar acesso':
                acesso = 'Request Access'
                if pd.isna(row['Empresa']) != 'Informação para criar acesso':
                    parceira = 'yes'
                    descricao_terceiro_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                else:
                    parceira = 'no'
                    descricao_user_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                    
            elif pd.isna(row['Tipo de solicitação do VPN']) == '2º Token':
                acesso = '2nd copy of Token'
                if pd.isna(row['Empresa']) != 'Informação para criar acesso':
                    parceira = 'yes'
                    descricao_terceiro_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                else:
                    parceira = 'no'
                    descricao_user_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                    
            elif pd.isna(row['Tipo de solicitação do VPN']) == 'Retirar acesso':
                acesso = 'Revoke Access'
                if pd.isna(row['Empresa']) != 'Informação para criar acesso':
                    parceira = 'yes'
                    descricao_terceiro_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
                else:
                    parceira = 'no'
                    descricao_user_vpn = f"Matricula: {matricula}\nEmpresa: {empresa} \nCargo: {cargo}\n{descricao}"
               

            chamado = janela_principal_vpn(nome_usuario_vpn, matricula, empresa, acesso_vpn, terceiro_vpn, email_terceiro_vpn, descricao_terceiro_vpn, descricao_user_vpn, atendimento_posto_vpn, localidade_vpn, perfil_solicitado_vpn, nivel_hierarquico_vpn, gerente_colab_vpn)    
            
        if chamado:
            df.at[index, 'Chamado'] = chamado
            df.to_excel(arquivo_excel, index = False, sheet_name='Chamado')
            print(f"Chamado criado: {chamado}")
            
        # Retornar para a página inicial
        time.sleep(8)
        navegador.get(pagina)
        time.sleep(3)

        #navegador.back()
        #time.sleep(10)
        if chamados_criados:
            print('\nTabela de chamados criados:')
            print(tabulate(chamados_criados, headers=['Nome', 'tipo de acesso','Numero do chamado'], tablefmt = 'pretty'))
        else:
            print("Nenhum chamado criado")

if __name__ == "__main__":
    main()